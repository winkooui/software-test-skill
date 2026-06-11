#!/usr/bin/env python3
"""Generate demo test cases from a PRD-like markdown file.

The script is intentionally deterministic: it provides a runnable example for
the skill repository without requiring an LLM or network access.
"""

from __future__ import annotations

import argparse
import csv
from pathlib import Path


HEADERS = [
    "id",
    "module",
    "function",
    "type",
    "priority",
    "name",
    "precondition",
    "steps",
    "test_data",
    "data_prep_method",
    "expected_result",
    "actual_result",
    "target_env",
    "env_constraint",
    "predecessor_case",
    "successor_case",
    "related_requirement",
    "related_test_point",
    "user_perspective_tag",
    "user_persona",
    "is_regression",
    "regression_trigger",
    "proximity_level",
    "regression_depth",
    "prod_gate",
    "exec_status",
    "executor",
    "exec_date",
    "bug_id",
    "notes",
]


def detect_features(prd_text: str) -> list[str]:
    feature_keywords = [
        ("账号密码登录", ["登录", "密码", "手机号", "邮箱"]),
        ("错误次数锁定", ["锁定", "错误次数", "15 分钟", "ACCOUNT_LOCKED"]),
        ("跳转恢复", ["redirect_url", "跳转", "受保护页面"]),
        ("登录安全", ["审计", "日志", "明文", "错误码"]),
        ("移动端体验", ["移动端", "375px", "防重复提交"]),
    ]
    features = []
    for feature, keywords in feature_keywords:
        if any(keyword in prd_text for keyword in keywords):
            features.append(feature)
    return features or ["核心业务流程"]


def build_cases(prd_text: str) -> list[dict[str, str]]:
    features = detect_features(prd_text)
    cases: list[dict[str, str]] = []

    case_templates = [
        {
            "function": "账号密码登录",
            "type": "功能测试",
            "suffix": "F",
            "priority": "P0",
            "name": "正确账号密码登录成功",
            "steps": "1. 打开登录页; 2. 输入正确账号和密码; 3. 点击登录",
            "expected_result": "登录成功，创建会话并进入目标页面",
            "test_data": "account=tester@example.com; password=ValidPass123",
            "prod_gate": "true",
        },
        {
            "function": "账号密码登录",
            "type": "安全测试",
            "suffix": "S",
            "priority": "P0",
            "name": "错误密码不暴露账号存在性",
            "steps": "1. 输入存在账号和错误密码; 2. 输入不存在账号和任意密码; 3. 对比错误提示",
            "expected_result": "两种失败场景返回一致提示，不泄露账号是否存在",
            "test_data": "existing=tester@example.com; missing=none@example.com",
            "prod_gate": "true",
        },
        {
            "function": "错误次数锁定",
            "type": "回归测试",
            "suffix": "R",
            "priority": "P0",
            "name": "连续5次错误后账号锁定",
            "steps": "1. 连续5次输入错误密码; 2. 第6次输入正确密码",
            "expected_result": "账号被锁定，返回ACCOUNT_LOCKED，锁定期间拒绝登录",
            "test_data": "account=lock@example.com; wrong=BadPass000",
            "prod_gate": "true",
        },
        {
            "function": "跳转恢复",
            "type": "安全测试",
            "suffix": "S",
            "priority": "P0",
            "name": "拒绝非本站redirect_url",
            "steps": "1. 构造外部redirect_url; 2. 完成登录; 3. 观察跳转目标",
            "expected_result": "系统拒绝外部URL，进入默认首页或安全提示页",
            "test_data": "redirect_url=https://evil.example",
            "prod_gate": "true",
        },
        {
            "function": "移动端体验",
            "type": "用户视角测试",
            "suffix": "N",
            "priority": "P2",
            "name": "移动端登录表单可完整操作",
            "steps": "1. 使用375px视口打开登录页; 2. 输入账号密码; 3. 查看按钮和错误提示布局",
            "expected_result": "输入框、按钮、错误提示无遮挡且可点击",
            "test_data": "viewport=375x812",
            "prod_gate": "false",
        },
    ]

    selected = [
        template
        for template in case_templates
        if template["function"] in features or template["function"] in {"登录安全", "账号密码登录"}
    ]
    if not selected:
        selected = case_templates[:3]

    for index, template in enumerate(selected, start=1):
        is_regression = template["type"] == "回归测试" or template["priority"] == "P0"
        is_user = template["type"] == "用户视角测试"
        cases.append(
            {
                "id": f"TC-AUTH-{index:02d}-{index:03d}-{template['suffix']}",
                "module": "认证登录",
                "function": template["function"],
                "type": template["type"],
                "priority": template["priority"],
                "name": template["name"],
                "precondition": "测试环境已部署，测试账号和依赖服务可用",
                "steps": template["steps"],
                "test_data": template["test_data"],
                "data_prep_method": "手工创建" if "account" in template["test_data"] else "无需准备",
                "expected_result": template["expected_result"],
                "actual_result": "",
                "target_env": "沙箱, SIT, UAT" if template["priority"] != "P2" else "沙箱, UAT",
                "env_constraint": "无",
                "predecessor_case": "",
                "successor_case": "",
                "related_requirement": f"PRD-AUTH-{index:03d}",
                "related_test_point": template["function"],
                "user_perspective_tag": "新手探索" if is_user else "",
                "user_persona": "移动端用户" if is_user else "",
                "is_regression": str(is_regression).lower(),
                "regression_trigger": "登录链路变更" if is_regression else "",
                "proximity_level": "高" if is_regression else "",
                "regression_depth": "全量" if is_regression else "",
                "prod_gate": template["prod_gate"],
                "exec_status": "未执行",
                "executor": "",
                "exec_date": "",
                "bug_id": "",
                "notes": "脚本示例生成，可由 skill 进一步扩展为完整用例集",
            }
        )
    return cases


def write_csv(path: Path, rows: list[dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as file:
        writer = csv.DictWriter(file, fieldnames=HEADERS)
        writer.writeheader()
        writer.writerows(rows)


def write_markdown(path: Path, rows: list[dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as file:
        file.write("# 测试用例生成结果\n\n")
        file.write(f"- 用例数: {len(rows)}\n")
        file.write("- 输出说明: 该文件由本仓库示例脚本生成，用于快速预览 skill 的交付格式。\n\n")
        file.write("| 用例编号 | 模块 | 类型 | 优先级 | 用例名称 | Prod门禁 |\n")
        file.write("|----------|------|------|--------|----------|----------|\n")
        for row in rows:
            file.write(
                f"| {row['id']} | {row['module']} | {row['type']} | {row['priority']} | "
                f"{row['name']} | {row['prod_gate']} |\n"
            )


def write_xlsx_if_available(path: Path, rows: list[dict[str, str]]) -> bool:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        return False

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "测试用例主表"
    sheet.append(HEADERS)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F2937")
    for row in rows:
        sheet.append([row.get(header, "") for header in HEADERS])
    sheet.freeze_panes = "A2"
    workbook.save(path)
    return True


def main() -> int:
    parser = argparse.ArgumentParser(description="Generate demo test cases from a PRD markdown file.")
    parser.add_argument("--input", default="examples/prd-login.md", help="Path to PRD markdown file.")
    parser.add_argument("--out-dir", default="test-output", help="Directory for generated artifacts.")
    parser.add_argument("--project", default="login", help="Project slug used in output filenames.")
    args = parser.parse_args()

    prd_path = Path(args.input)
    prd_text = prd_path.read_text(encoding="utf-8")
    rows = build_cases(prd_text)
    out_dir = Path(args.out_dir)

    csv_path = out_dir / f"{args.project}-test-cases.csv"
    md_path = out_dir / f"{args.project}-test-cases.md"
    xlsx_path = out_dir / f"{args.project}-test-cases.xlsx"
    write_csv(csv_path, rows)
    write_markdown(md_path, rows)
    xlsx_created = write_xlsx_if_available(xlsx_path, rows)

    print(f"Generated {len(rows)} test cases")
    print(f"- {csv_path}")
    print(f"- {md_path}")
    if xlsx_created:
        print(f"- {xlsx_path}")
    else:
        print("- Excel skipped: install openpyxl to enable .xlsx output")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
