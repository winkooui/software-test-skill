#!/usr/bin/env python3
"""Generate a lightweight test report from execution records CSV."""

from __future__ import annotations

import argparse
import csv
from collections import Counter, defaultdict
from pathlib import Path


STATUSES = ["通过", "失败", "阻塞", "跳过", "未执行"]


def read_records(path: Path) -> list[dict[str, str]]:
    with path.open(newline="", encoding="utf-8") as file:
        return list(csv.DictReader(file))


def percent(part: int, total: int) -> str:
    return "0.0%" if total == 0 else f"{part / total * 100:.1f}%"


def readiness_score(records: list[dict[str, str]]) -> tuple[int, list[tuple[str, str, str]]]:
    score = 100
    risks: list[tuple[str, str, str]] = []
    p0_bad = [
        row
        for row in records
        if row.get("priority") == "P0" and row.get("status") not in {"通过", "跳过"}
    ]
    fatal_bugs = [row for row in records if row.get("bug_severity") == "致命"]
    severe_bugs = [row for row in records if row.get("bug_severity") == "严重"]
    blocked = [row for row in records if row.get("status") == "阻塞"]

    score -= len(p0_bad) * 3
    score -= len(fatal_bugs) * 10
    score -= len(severe_bugs) * 3
    score -= len(blocked) * 2

    if p0_bad:
        risks.append(("高", f"P0 用例未全部通过，共 {len(p0_bad)} 条", "修复后重新执行 P0 回归"))
    if fatal_bugs:
        bug_ids = ", ".join(row.get("bug_id", "-") for row in fatal_bugs)
        risks.append(("高", f"存在致命缺陷: {bug_ids}", "上线前必须关闭致命缺陷"))
    if severe_bugs:
        bug_ids = ", ".join(row.get("bug_id", "-") for row in severe_bugs)
        risks.append(("中", f"存在严重缺陷: {bug_ids}", "评估影响面并完成重点回归"))
    if blocked:
        risks.append(("中", f"存在阻塞用例，共 {len(blocked)} 条", "补齐环境或数据后复测"))

    return max(0, min(100, score)), risks


def summarize_by_module(records: list[dict[str, str]]) -> dict[str, Counter]:
    summary: dict[str, Counter] = defaultdict(Counter)
    for row in records:
        module = row.get("module") or "未分类"
        summary[module][row.get("status") or "未执行"] += 1
    return summary


def conclusion(score: int, risks: list[tuple[str, str, str]]) -> str:
    has_high = any(level == "高" for level, _, _ in risks)
    if score >= 90 and not has_high:
        return "通过"
    if score >= 75:
        return "有条件通过"
    return "不通过"


def write_markdown(path: Path, records: list[dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    counts = Counter(row.get("status") or "未执行" for row in records)
    total = len(records)
    passed = counts["通过"]
    score, risks = readiness_score(records)
    result = conclusion(score, risks)
    module_summary = summarize_by_module(records)

    with path.open("w", encoding="utf-8") as file:
        file.write("# 测试报告\n\n")
        file.write("## KPI 概览\n\n")
        file.write("| 指标 | 数值 |\n|------|------|\n")
        file.write(f"| 总用例数 | {total} |\n")
        file.write(f"| 已执行 | {total - counts['未执行'] - counts['跳过']} |\n")
        for status in STATUSES:
            file.write(f"| {status} | {counts[status]} |\n")
        file.write(f"| 通过率 | {percent(passed, total)} |\n")
        file.write(f"| 上线就绪度 | {score} |\n")
        file.write(f"| 结论 | {result} |\n\n")

        file.write("## 主要风险\n\n")
        if risks:
            file.write("| 等级 | 风险 | 建议 |\n|------|------|------|\n")
            for level, risk, action in risks:
                file.write(f"| {level} | {risk} | {action} |\n")
        else:
            file.write("未识别到阻塞上线的主要风险。\n")

        file.write("\n## 模块统计\n\n")
        file.write("| 模块 | 总数 | 通过 | 失败 | 阻塞 | 跳过 | 通过率 |\n")
        file.write("|------|------|------|------|------|------|--------|\n")
        for module, counter in sorted(module_summary.items()):
            module_total = sum(counter.values())
            file.write(
                f"| {module} | {module_total} | {counter['通过']} | {counter['失败']} | "
                f"{counter['阻塞']} | {counter['跳过']} | {percent(counter['通过'], module_total)} |\n"
            )


def write_summary_csv(path: Path, records: list[dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    module_summary = summarize_by_module(records)
    with path.open("w", newline="", encoding="utf-8") as file:
        writer = csv.writer(file)
        writer.writerow(["module", "total", "passed", "failed", "blocked", "skipped", "pass_rate"])
        for module, counter in sorted(module_summary.items()):
            total = sum(counter.values())
            writer.writerow(
                [
                    module,
                    total,
                    counter["通过"],
                    counter["失败"],
                    counter["阻塞"],
                    counter["跳过"],
                    percent(counter["通过"], total),
                ]
            )


def write_xlsx_if_available(path: Path, records: list[dict[str, str]]) -> bool:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        return False

    path.parent.mkdir(parents=True, exist_ok=True)
    counts = Counter(row.get("status") or "未执行" for row in records)
    score, risks = readiness_score(records)
    workbook = Workbook()
    dashboard = workbook.active
    dashboard.title = "Dashboard"
    dashboard.append(["指标", "数值"])
    for cell in dashboard[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F2937")
    dashboard.append(["总用例数", len(records)])
    for status in STATUSES:
        dashboard.append([status, counts[status]])
    dashboard.append(["上线就绪度", score])
    dashboard.append(["结论", conclusion(score, risks)])

    detail = workbook.create_sheet("执行明细")
    if records:
        headers = list(records[0].keys())
        detail.append(headers)
        for row in records:
            detail.append([row.get(header, "") for header in headers])

    risk_sheet = workbook.create_sheet("风险评估")
    risk_sheet.append(["等级", "风险", "建议"])
    for risk in risks:
        risk_sheet.append(list(risk))

    workbook.save(path)
    return True


def main() -> int:
    parser = argparse.ArgumentParser(description="Generate a test report from execution records CSV.")
    parser.add_argument("--input", default="examples/execution-records.csv", help="Path to execution CSV.")
    parser.add_argument("--out-dir", default="test-output", help="Directory for generated artifacts.")
    parser.add_argument("--project", default="login", help="Project slug used in output filenames.")
    args = parser.parse_args()

    records = read_records(Path(args.input))
    out_dir = Path(args.out_dir)
    md_path = out_dir / f"{args.project}-test-report.md"
    csv_path = out_dir / f"{args.project}-report-summary.csv"
    xlsx_path = out_dir / f"{args.project}-test-report.xlsx"
    write_markdown(md_path, records)
    write_summary_csv(csv_path, records)
    xlsx_created = write_xlsx_if_available(xlsx_path, records)

    print(f"Generated report from {len(records)} execution records")
    print(f"- {md_path}")
    print(f"- {csv_path}")
    if xlsx_created:
        print(f"- {xlsx_path}")
    else:
        print("- Excel skipped: install openpyxl to enable .xlsx output")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
